from openpyxl import load_workbook, Workbook
import cryptocode
import os


# Clase del scaner del libro de excel. El libro debe estar en el mismo
# directorio que el módulo.

class Lector:
    # Constructor de "Lector". Necesita el nombre del archivo excel del que sacar los datos.
    def __init__(self, name, ruta):
        self.name = name
        self.wb = load_workbook(ruta, data_only=True)
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.intermedia = load_workbook('assets/tabla-intermedia.xlsx')
        self.andina_codesheet = self.intermedia['Andina']
        self.od_codesheet = self.intermedia['Oscar-David']
        self.ser_codesheet = self.intermedia['La-Serenisima']
        self.bees_codesheet = self.intermedia['Bees']
        self.sku_list = []
        self.name_list = []
        self.prov_list = []
        self.providers = []
        self.code_list = []
        self.date_list = []
        self.datos = {}
    
    def obtener_proveedores(self):
        self.obtener_proveedor()
        for prov in self.prov_list:
            if ',' not in prov:
                if prov not in self.providers:
                    self.providers.append(prov)
            else:
                elements = prov.split(',')
                for element in elements:
                    if element not in self.providers:
                        self.providers.append(element)
        return self.providers

    # Método que toma los SKU's de la primer fila del excel.
    def obtener_skus(self):
        for column_data in self.ws['A']:
            if column_data.value != None:
                valor = str(column_data.value)
                self.sku_list.append(valor)
        return self.sku_list
    
    # Método que toma los nombres de los artículos de la tercer columna del excel.
    def obtener_nombres(self):
        for column_data in self.ws['C']:
            if column_data.value != None:
                valor = str(column_data.value)
                self.name_list.append(valor)
        return self.name_list
    
    # Método que toma los códigos de los artículos de la segunda columna del excel.
    def obtener_codigo(self):
        for column_data in self.ws['B']:
            if column_data.value != None:
                valor = str(column_data.value)
                self.code_list.append(valor)
        return self.code_list
    
    # Método que toma los nombres de los proveedores de los artículos de la séptima columna del excel.
    def obtener_proveedor(self):
        for column_data in self.ws['G']:
            if column_data.value != None:
                valor = str(column_data.value)
                self.prov_list.append(valor)
        return self.prov_list
    
    def obtener_fechas(self):
        for column_data in self.ws['H']:
            if column_data.value != None:
                valor = str(column_data.value)
                self.date_list.append(valor)
        return self.date_list

    # Método que toma los datos utilizando los métodos anteriores
    def obtener_datos(self):
        articulos = []
        self.obtener_skus()
        self.obtener_codigo()
        self.obtener_nombres()
        self.obtener_proveedor()
        self.obtener_fechas()
        self.codigo_fecha = {}
        for code in range(len(self.code_list)):
            articulo = {
                'nombre': f'{self.name_list[code]}',
                'SKU': f'{self.sku_list[code]}',
                'fecha': f'{self.date_list[code]}',
                'codigo': f'{self.code_list[code]}',
                'proveedor': f'{self.prov_list[code]}',
            }
            if self.prov_list[code] == 'DISTRIBUIDORA ANDINA' or self.prov_list[code] == 'ANDINA SRL':
                articulo['proveedor'] = 'ANDINA'
            elif self.prov_list[code] == 'OSCAR DAVID MAYORISTA':
                articulo['proveedor'] = 'OSCAR DAVID'
            elif self.prov_list[code] == 'SANTO GUILIANO' or self.prov_list[code] == 'ESTEBAN PANELLA':
                articulo['proveedor'] = 'BEES'
            articulos.append(articulo)
        
        return articulos

    def intercode(self, articulos):
        self.and_codes = {}
        self.od_codes = {}
        self.ser_codes = {}
        self.bees_codes = {}

        for row in self.andina_codesheet.rows:
            if row[3].value != 'Andina':
                self.and_codes[str(int(row[1].value))] = str(int(row[3].value)) 
        for row in self.od_codesheet.rows:
            if row[3].value != 'Oscar-David':
                self.od_codes[str(int(row[1].value))] = str(int(row[3].value))
        for row in self.ser_codesheet.rows:
            if row[3].value != 'La-Serenisima':
                self.ser_codes[str(int(row[1].value))] = str(int(row[3].value))
        for row in self.bees_codesheet.rows:
            if row[3].value != 'Bees':
                self.bees_codes[str(int(row[1].value))] = str(int(row[3].value))

        
        for articulo in articulos:
            if 'ANDINA' in articulo.proveedor.nombre:
                try:
                    articulo.cod_externo = self.and_codes[articulo.codigo]
                except:
                    pass
            elif 'OSCAR DAVID' in articulo.proveedor.nombre:
                try:
                    articulo.cod_externo = self.od_codes[articulo.codigo]
                except:
                    pass
            elif 'SERENISIMA' in articulo.proveedor.nombre:
                try:
                    articulo.cod_externo = self.ser_codes[articulo.codigo]
                except:
                    pass
            elif 'SANTO GUILIANO' in articulo.proveedor.nombre or 'JOSE ESTEBAN PANELLA' in articulo.proveedor.nombre or 'BEES' in articulo.proveedor.nombre:
                try:
                    articulo.cod_externo = self.bees_codes[articulo.codigo]
                except:
                    pass
        
        return articulos

    

    # Método que actualiza los precios en un excel actualizado
    def actualizar_precios(self, articulos):
        self.nombre_archivo = f'{self.name}-ACTUALIZADO.xlsx'
        wb_act = Workbook()
        ws_act = wb_act[wb_act.sheetnames[0]]
        i = 1
        ws_act.cell(1, 1, 'Il Porco')
        ws_act.cell(1, 2, 'Artículo')
        ws_act.cell(1, 3, 'SKU')
        ws_act.cell(1, 4, 'Distribuidor')
        ws_act.cell(1, 5, 'Cód. Externo')
        ws_act.cell(1, 6, 'Nuevo Precio')
        ws_act.cell(1, 7, 'Última fecha')

        for articulo in articulos:
            ws_act.cell(i+1, 1, articulo.codigo)
            ws_act.cell(i+1, 2, articulo.nombre)
            ws_act.cell(i+1, 3, articulo.sku)
            ws_act.cell(i+1, 4, articulo.proveedor.nombre)
            ws_act.cell(i+1, 5, articulo.cod_externo)
            if articulo.precio != 'Sin precio' and articulo.proveedor.nombre == 'MAXICONSUMO' or articulo.precio != 'Sin precio' and articulo.proveedor.nombre == 'ANDINA':
                ws_act.cell(i+1, 6, articulo.precio[2:])
            else:
                ws_act.cell(i+1, 6, articulo.precio)
            ws_act.cell(i+1, 7, articulo.fecha)
            i+=1
        
        wb_act.save(f'{os.getcwd()}/archivos/{self.nombre_archivo}')

    def abrir_planilla_resultado(self):
        os.system(f'start excel.exe "{os.getcwd()}\\archivos\\{self.nombre_archivo}"')

